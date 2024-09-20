import pandas as pd
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

def find_common_data(file, sheet1, sheet2, column_name):
    try:
        # 使用 openpyxl 加载工作簿以检查工作表是否存在
        wb = load_workbook(file)
        if sheet1 not in wb.sheetnames or sheet2 not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet1}' 或 '{sheet2}' 不存在于文件中")

        # 读取Excel文件的两个工作表
        df1 = pd.read_excel(file, sheet_name=sheet1)
        df2 = pd.read_excel(file, sheet_name=sheet2)
        
        # 检查列名是否存在于两个数据框中
        if column_name not in df1.columns:
            raise ValueError(f"列 '{column_name}' 在工作表 '{sheet1}' 中不存在")
        if column_name not in df2.columns:
            raise ValueError(f"列 '{column_name}' 在工作表 '{sheet2}' 中不存在")
        
        # 找出指定列中相同的数据，并保留所有列
        common_data = pd.merge(df1, df2, on=column_name, how='inner', suffixes=('_'+sheet1, '_'+sheet2))
        
        # 生成带有时间戳的新文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'common_data_{timestamp}.xlsx'
        
        # 创建 data 文件夹（如果不存在）
        data_folder = 'data'
        os.makedirs(data_folder, exist_ok=True)
        
        # 将输出文件路径设置在 data 文件夹下
        output_path = os.path.join(data_folder, output_file)
        
        # 导出结果到新的Excel文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            common_data.to_excel(writer, index=False, sheet_name='Common Data')
            
            # 获取新工作簿和工作表
            wb_new = writer.book
            ws_new = wb_new['Common Data']
            
            # 复制原始工作表的字体样式
            for col in range(1, ws_new.max_column + 1):
                original_cell = wb[sheet1].cell(row=1, column=col)
                new_cell = ws_new.cell(row=1, column=col)
                new_cell.font = Font(
                    name=original_cell.font.name,
                    size=original_cell.font.size,
                    bold=original_cell.font.bold,
                    italic=original_cell.font.italic,
                    color=original_cell.font.color
                )

        print(f"相同的数据已导出到 {os.path.abspath(output_path)}")
        print(f"共找到 {len(common_data)} 条相同的数据")
    
    except FileNotFoundError:
        print(f"错误：找不到文件 '{file}'")
    except ValueError as e:
        print(f"错误：{str(e)}")
    except PermissionError:
        print(f"错误：没有权限创建或写入文件。请确保您有足够的权限，并且文件没有被其他程序占用。")
    except Exception as e:
        print(f"发生未知错误：{str(e)}")

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("使用方法: python find_common_data.py <file.xlsx> <sheet1> <sheet2> <column_name>")
        sys.exit(1)
    
    file = sys.argv[1]
    sheet1 = sys.argv[2]
    sheet2 = sys.argv[3]
    column_name = sys.argv[4]
    
    find_common_data(file, sheet1, sheet2, column_name)